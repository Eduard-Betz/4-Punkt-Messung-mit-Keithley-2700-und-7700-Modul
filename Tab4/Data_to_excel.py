# Datei: Data_to_excel.py

from global_vars import TKBoardVariabeln, TK_Fehler, PlotAuswahl  # TK_Fehler importieren

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl.utils
import tkinter as tk
from tkinter import filedialog

# Debug-Flag definieren
debug_Tab4 = False  # Debug-Ausgaben sind standardmäßig deaktiviert

# Debug-Ausgabefunktion
def debug_print(*args, **kwargs):
    if debug_Tab4:
        print(*args, **kwargs)

def print_tk_data():
    # Öffne einen Dialog, um den Speicherort der Excel-Datei zu wählen
    file_path = filedialog.asksaveasfilename(
        defaultextension='.xlsx',
        filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
        title='Excel-Datei speichern'
    )

    if not file_path:
        # Benutzer hat den Dialog abgebrochen
        return

    data = TKBoardVariabeln.board_variablen

    # Debug-Ausgabe hinzufügen
    debug_print("Inhalt von TKBoardVariabeln.board_variablen:")
    debug_print(data)

    # Listen zur Speicherung der Daten für die Tabellen
    normal_data_list = []
    avg_data_list = []

    # Daten extrahieren
    for board_nr, resistors in data.items():
        for resistor_nr, values in resistors.items():
            # Gemeinsame Daten
            row_common = {
                'Board Nr.': board_nr,
                'Probe': '',  # Diese Spalte bleibt leer
            }

            # R²-Werte und delta_alpha_total aus TK_Fehler holen
            fehler_data = TK_Fehler.get_board_data(board_nr)
            resistor_fehler = fehler_data.get(resistor_nr, {}) if fehler_data else {}

            # Normal Daten
            if 'normal' in values:
                normal = values['normal']
                # Runden der Temperaturwerte auf zwei Dezimalstellen
                temp_steigend = normal.get('temp_bereich_steigend', (None, None))
                temp_fallend = normal.get('temp_bereich_sinkend', (None, None))

                temp_steigend = tuple(round(t, 2) if isinstance(t, (int, float)) else t for t in temp_steigend)
                temp_fallend = tuple(round(t, 2) if isinstance(t, (int, float)) else t for t in temp_fallend)

                # Fehlerdaten für normale Daten
                normal_fehler = resistor_fehler.get('normal', {})
                r_squared_steigend = normal_fehler.get('r_squared_steigend', '')
                r_squared_fallend = normal_fehler.get('r_squared_sinkend', '')
                delta_alpha_total_steigend = normal_fehler.get('delta_alpha_total_steigend', '')
                delta_alpha_total_fallend = normal_fehler.get('delta_alpha_total_sinkend', '')

                normal_row = row_common.copy()
                normal_row.update({
                    'TK_steigende': normal.get('steigung_steigend', ''),
                    'Fehler ±_steigende': delta_alpha_total_steigend,  # ∆αGesamt hinzufügen
                    'R²_steigende': r_squared_steigend,  # R²-Wert hinzufügen
                    'min temp_steigende': temp_steigend[0],
                    'max temp_steigende': temp_steigend[1],

                    'TK_fallende': normal.get('steigung_sinkend', ''),
                    'Fehler ±_fallende': delta_alpha_total_fallend,  # ∆αGesamt hinzufügen
                    'R²_fallende': r_squared_fallend,     # R²-Wert hinzufügen
                    'min temp_fallende': temp_fallend[0],
                    'max temp_fallende': temp_fallend[1],
                })
                normal_data_list.append(normal_row)

            # Avg Daten
            if 'avg' in values:
                avg = values['avg']
                # Runden der Temperaturwerte auf zwei Dezimalstellen
                temp_steigend = avg.get('temp_bereich_steigend', (None, None))
                temp_fallend = avg.get('temp_bereich_sinkend', (None, None))

                temp_steigend = tuple(round(t, 2) if isinstance(t, (int, float)) else t for t in temp_steigend)
                temp_fallend = tuple(round(t, 2) if isinstance(t, (int, float)) else t for t in temp_fallend)

                # Fehlerdaten für avg Daten
                avg_fehler = resistor_fehler.get('avg', {})
                r_squared_steigend_avg = avg_fehler.get('r_squared_steigend', '')
                r_squared_fallend_avg = avg_fehler.get('r_squared_sinkend', '')
                delta_alpha_total_steigend_avg = avg_fehler.get('delta_alpha_total_steigend', '')
                delta_alpha_total_fallend_avg = avg_fehler.get('delta_alpha_total_sinkend', '')

                avg_row = row_common.copy()
                avg_row.update({
                    'TK_steigende': avg.get('steigung_steigend', ''),
                    'Fehler ±_steigende': delta_alpha_total_steigend_avg,  # ∆αGesamt hinzufügen
                    'R²_steigende': r_squared_steigend_avg,  # R²-Wert hinzufügen
                    'min temp_steigende': temp_steigend[0],
                    'max temp_steigende': temp_steigend[1],

                    'TK_fallende': avg.get('steigung_sinkend', ''),
                    'Fehler ±_fallende': delta_alpha_total_fallend_avg,  # ∆αGesamt hinzufügen
                    'R²_fallende': r_squared_fallend_avg,     # R²-Wert hinzufügen
                    'min temp_fallende': temp_fallend[0],
                    'max temp_fallende': temp_fallend[1],
                })
                avg_data_list.append(avg_row)

    # Überprüfen, ob Daten in den Listen vorhanden sind
    if not normal_data_list and not avg_data_list:
        debug_print("Keine Daten zum Exportieren gefunden.")
        return

    # Spalten definieren und neue Spalten hinzufügen
    columns = [
        'Board Nr.', 'Probe',
        'TK_steigende', 'Fehler ±_steigende', 'R²_steigende', 'min temp_steigende', 'max temp_steigende',
        'TK_fallende', 'Fehler ±_fallende', 'R²_fallende', 'min temp_fallende', 'max temp_fallende'
    ]

    normal_df = pd.DataFrame(normal_data_list, columns=columns)
    avg_df = pd.DataFrame(avg_data_list, columns=columns)

    # Daten in Excel schreiben
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        normal_header = "TK mit Board Temperatur"
        avg_header = "TK mit gemittelte Temperatur über alle Boards"

        # Schreibe die normale Tabelle ab Zeile 3 (startrow=2 wegen nullbasierter Indizierung)
        normal_table_start_row = 3
        normal_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=normal_table_start_row - 1, header=True)

        # Jetzt haben wir Daten geschrieben, wir können die workbook und worksheet erhalten
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Schreibe die Überschrift für die erste Tabelle in Zeile 1
        normal_header_row_excel = 1
        worksheet.cell(row=normal_header_row_excel, column=1).value = normal_header
        worksheet.merge_cells(start_row=normal_header_row_excel, start_column=1, end_row=normal_header_row_excel, end_column=len(columns))
        header_cell = worksheet.cell(row=normal_header_row_excel, column=1)
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal='center')

        # Füge die Gruppenkopfzeile hinzu
        group_header_row = normal_table_start_row - 1  # Zeile für Gruppenkopfzeilen
        column_header_row = normal_table_start_row     # Zeile für Spaltenüberschriften

        # Gruppenkopfzeilen hinzufügen und Zellen zusammenführen
        # "Steigende Flanke" über Spalten 3 bis 7
        worksheet.merge_cells(start_row=group_header_row, start_column=3, end_row=group_header_row, end_column=7)
        steigende_cell = worksheet.cell(row=group_header_row, column=3)
        steigende_cell.value = "Steigende Flanke"
        steigende_cell.alignment = Alignment(horizontal='center', vertical='center')
        steigende_cell.font = Font(bold=True)
        steigende_cell.border = Border(
            bottom=Side(style='thick'),  # Ändere von 'medium' zu 'thick'
            top=Side(style='thick'),     # Ändere von 'medium' zu 'thick'
            left=Side(style='thick'),    # Ändere von 'medium' zu 'thick'
            right=Side(style='thick')    # Ändere von 'medium' zu 'thick'
        )

        # "Fallende Flanke" über Spalten 8 bis 12
        worksheet.merge_cells(start_row=group_header_row, start_column=8, end_row=group_header_row, end_column=12)
        fallende_cell = worksheet.cell(row=group_header_row, column=8)
        fallende_cell.value = "Fallende Flanke"
        fallende_cell.alignment = Alignment(horizontal='center', vertical='center')
        fallende_cell.font = Font(bold=True)
        fallende_cell.border = Border(
            bottom=Side(style='thick'),  # Ändere von 'medium' zu 'thick'
            top=Side(style='thick'),     # Ändere von 'medium' zu 'thick'
            left=Side(style='thick'),    # Ändere von 'medium' zu 'thick'
            right=Side(style='thick')    # Ändere von 'medium' zu 'thick'
        )


        # Suffixe aus Spaltenüberschriften entfernen und '∆αGesamt' zu 'Fehler ±' ersetzen
        for col_num in range(1, len(columns) + 1):
            cell = worksheet.cell(row=column_header_row, column=col_num)
            if '_steigende' in cell.value:
                cell.value = cell.value.replace('_steigende', '')
            elif '_fallende' in cell.value:
                cell.value = cell.value.replace('_fallende', '')
            # Ersetzen von '∆αGesamt' durch 'Fehler ±'
            if '∆αGesamt' in cell.value:
                cell.value = 'Fehler ±'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            # Rahmen für Spaltenüberschriften
            cell.border = Border(bottom=Side(style='thin'))


        # Berechne die Endzeile der normalen Tabelle
        normal_table_end_row = normal_table_start_row + len(normal_df)  # Ohne -1

        # Füge 10 leere Zeilen zwischen den Tabellen
        avg_header_row = normal_table_end_row + 10

        # Schreibe die Überschrift für die zweite Tabelle
        worksheet.cell(row=avg_header_row, column=1).value = avg_header
        worksheet.merge_cells(start_row=avg_header_row, start_column=1, end_row=avg_header_row, end_column=len(columns))
        header_cell = worksheet.cell(row=avg_header_row, column=1)
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal='center')

        # Schreibe die avg-Tabelle ab der berechneten Zeile
        avg_table_start_row = avg_header_row + 2  # +2 um die Gruppenkopfzeilen einzuschließen
        avg_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=avg_table_start_row - 1, header=True)

        # Füge die Gruppenkopfzeile für die avg-Tabelle hinzu
        group_header_row_avg = avg_table_start_row - 1  # Zeile für Gruppenkopfzeilen
        column_header_row_avg = avg_table_start_row     # Zeile für Spaltenüberschriften

        # Gruppenkopfzeilen hinzufügen und Zellen zusammenführen
        # "Steigende Flanke" über Spalten 3 bis 7
        worksheet.merge_cells(start_row=group_header_row_avg, start_column=3, end_row=group_header_row_avg, end_column=7)
        steigende_cell_avg = worksheet.cell(row=group_header_row_avg, column=3)
        steigende_cell_avg.value = "Steigende Flanke"
        steigende_cell_avg.alignment = Alignment(horizontal='center', vertical='center')
        steigende_cell_avg.font = Font(bold=True)
        steigende_cell_avg.border = Border(
            bottom=Side(style='medium'),
            top=Side(style='medium'),
            left=Side(style='medium'),
            right=Side(style='medium')
        )

        # "Fallende Flanke" über Spalten 8 bis 12
        worksheet.merge_cells(start_row=group_header_row_avg, start_column=8, end_row=group_header_row_avg, end_column=12)
        fallende_cell_avg = worksheet.cell(row=group_header_row_avg, column=8)
        fallende_cell_avg.value = "Fallende Flanke"
        fallende_cell_avg.alignment = Alignment(horizontal='center', vertical='center')
        fallende_cell_avg.font = Font(bold=True)
        fallende_cell_avg.border = Border(
            bottom=Side(style='medium'),
            top=Side(style='medium'),
            left=Side(style='medium'),
            right=Side(style='medium')
        )

        # Suffixe aus Spaltenüberschriften entfernen und Formatierung anwenden
        for col_num in range(1, len(columns) + 1):
            cell = worksheet.cell(row=column_header_row_avg, column=col_num)
            if '_steigende' in cell.value:
                cell.value = cell.value.replace('_steigende', '')
            elif '_fallende' in cell.value:
                cell.value = cell.value.replace('_fallende', '')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            # Rahmen für Spaltenüberschriften
            cell.border = Border(bottom=Side(style='thin'))

        # Berechne die Endzeile der avg-Tabelle
        avg_table_end_row = avg_table_start_row + len(avg_df)  # Ohne -1

        # Rahmen und Formatierungen anwenden
        # Definiere Rahmenstile
        thin_border = Side(border_style="thin", color="000000")
        thick_border = Side(border_style="thick", color="000000")

        def apply_table_border(ws, start_row, end_row, start_col, end_col):
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                for cell in row:
                    border_sides = {
                        'left': thin_border,
                        'right': thin_border,
                        'top': thin_border,
                        'bottom': thin_border
                    }

                    # Äußere Ränder der Tabelle mit 'thick'
                    if cell.row == start_row:
                        border_sides['top'] = thick_border
                    if cell.row == end_row:
                        border_sides['bottom'] = thick_border
                    if cell.column == start_col:
                        border_sides['left'] = thick_border
                    if cell.column == end_col:
                        border_sides['right'] = thick_border

                    cell.border = Border(**border_sides)


        # Rahmen für die normale Tabelle anwenden (einschließlich Gruppenkopfzeilen)
        apply_table_border(
            worksheet,
            group_header_row,  # Startzeile des Rahmens (Gruppenkopfzeilen)
            normal_table_end_row,  # Endzeile der Daten
            1,
            len(columns)
        )

        # Rahmen für die avg-Tabelle anwenden (einschließlich Gruppenkopfzeilen)
        apply_table_border(
            worksheet,
            group_header_row_avg,  # Startzeile des Rahmens (Gruppenkopfzeilen)
            avg_table_end_row,     # Endzeile der Daten
            1,
            len(columns)
        )

        # Spaltenbreiten anpassen
        for i, column in enumerate(columns, 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

        # Ausrichtung für alle Zellen anpassen
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Erfolgsmeldung
    print(f"Excel-Datei wurde gespeichert unter: {file_path}")


def TK_auswahl_zu_excel():

    if PlotAuswahl is None:
        print("Die Variable PlotAuswahl ist nicht definiert.")
        return

    # Öffne ein Speichern-unter-Dialog, um den Dateinamen und Speicherort auszuwählen
    root = tk.Tk()
    root.withdraw()  # Versteckt das Hauptfenster
    root.call('wm', 'attributes', '.', '-topmost', True)  # Bringt das Fenster in den Vordergrund

    dateiname = filedialog.asksaveasfilename(
        defaultextension='.xlsx',
        filetypes=[('Excel Dateien', '*.xlsx'), ('Alle Dateien', '*.*')],
        title="Speichern unter"
    )

    if not dateiname:
        print("Speichern abgebrochen.")
        return

    # Erstelle den ExcelWriter mit 'openpyxl' Engine
    with pd.ExcelWriter(dateiname, engine='openpyxl') as writer:

        # Liste der Datensätze und zugehörigen Sheetnamen
        datasets = []
        if hasattr(PlotAuswahl, 'steigung') and PlotAuswahl.steigung is not None:
            datasets.append(('steigende Flanke', PlotAuswahl.steigung))
        if hasattr(PlotAuswahl, 'sinkend') and PlotAuswahl.sinkend is not None:
            datasets.append(('fallende Flanke', PlotAuswahl.sinkend))

        if not datasets:
            print("Keine gültigen Daten in PlotAuswahl gefunden.")
            return

        for sheet_name, plot_data in datasets:
            status, boards_data = plot_data

            # Ermitteln der maximalen Anzahl von Messpunkten
            max_messpunkte = max(len(data[1]) for data in boards_data.values())

            # Initialisiere die Spalten
            spalten = ['Messpunkt']

            # Erstelle ein Dictionary, das die Anzahl der Widerstände pro Board speichert
            widerstand_counts = {}

            # Boards sortieren für konsistente Spaltenreihenfolge
            sorted_boards = sorted(boards_data.keys())

            for board_name in sorted_boards:
                board_status, datenpunkte = boards_data[board_name]
                # Anzahl der Widerstände ermitteln (Annahme: Temperatur ist der letzte Wert, Zeitpunkt ist der erste)
                beispiel_datenpunkt = datenpunkte[0]
                anzahl_widerstaende = len(beispiel_datenpunkt) - 2  # - Zeitpunkt, - Temperatur
                widerstand_counts[board_name] = anzahl_widerstaende

            # Erstelle eine Liste von Dictionaries für jeden Messpunkt
            daten = []
            for messpunkt_index in range(max_messpunkte):
                messpunkt_daten = {'Messpunkt': messpunkt_index + 1}
                for board_name in sorted_boards:
                    board_status, datenpunkte = boards_data[board_name]
                    anzahl_widerstaende = widerstand_counts[board_name]
                    if messpunkt_index < len(datenpunkte):
                        datenpunkt = datenpunkte[messpunkt_index]
                        # Zeitpunkt ignorieren
                        widerstandswerte = datenpunkt[1:-1]
                        temperatur = datenpunkt[-1]
                        # Annahme: Nur ein Widerstand pro Board
                        messpunkt_daten[f'{board_name} Widerstand 1'] = widerstandswerte[0] if widerstandswerte else None
                        messpunkt_daten[f'{board_name} Temperatur'] = temperatur
                    else:
                        # Fehlende Daten mit None auffüllen
                        messpunkt_daten[f'{board_name} Widerstand 1'] = None
                        messpunkt_daten[f'{board_name} Temperatur'] = None
                daten.append(messpunkt_daten)

            # Erstelle den DataFrame
            df = pd.DataFrame(daten)

            # Schreibe den DataFrame in eine Excel-Datei mit Formatierungen
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=3)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Erste Zeile: Titel
            header_title = f"Messwerte mit Boardtemperatur {sheet_name}"
            total_columns = df.shape[1] + 1  # +1 für die zusätzliche Spalte
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
            cell = worksheet.cell(row=1, column=1)
            cell.value = header_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            # Zweite Zeile: 'Messpunkt', 'Board 1', 'Board 2', ..., 'Gemittelte Temperatur'
            worksheet.cell(row=2, column=1).value = 'Messpunkt'
            worksheet.cell(row=2, column=1).font = Font(bold=True)
            worksheet.cell(row=2, column=1).alignment = Alignment(horizontal='center')
            worksheet.cell(row=2, column=1).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            col = 2
            for board_name in sorted_boards:
                worksheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
                cell = worksheet.cell(row=2, column=col)
                cell.value = board_name
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                col += 2

            # Zusätzliche Spalte für 'Gemittelte Temperatur'
            cell = worksheet.cell(row=2, column=col)
            cell.value = 'Gemittelte Temperatur'
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            # Dritte Zeile: 'Widerstand 1', 'Temperatur', ..., ''
            worksheet.cell(row=3, column=1).value = ''
            worksheet.cell(row=3, column=1).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            col = 2
            for board_name in sorted_boards:
                cell_widerstand = worksheet.cell(row=3, column=col)
                cell_widerstand.value = 'Widerstand 1'
                cell_widerstand.font = Font(bold=True)
                cell_widerstand.alignment = Alignment(horizontal='center')
                cell_widerstand.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

                cell_temperatur = worksheet.cell(row=3, column=col + 1)
                cell_temperatur.value = 'Temperatur'
                cell_temperatur.font = Font(bold=True)
                cell_temperatur.alignment = Alignment(horizontal='center')
                cell_temperatur.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

                col += 2

            # Leere Zelle unter 'Gemittelte Temperatur'
            cell = worksheet.cell(row=3, column=col)
            cell.value = ''
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            # Schreibe die Daten und füge die Formel für die gemittelte Temperatur hinzu
            for row_idx, row in enumerate(df.itertuples(index=False), start=4):
                # Messpunkt
                worksheet.cell(row=row_idx, column=1).value = row[0]
                worksheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')

                col = 2
                temperatur_cells = []
                for board_name in sorted_boards:
                    # Widerstand
                    value_widerstand = row[df.columns.get_loc(f'{board_name} Widerstand 1')]
                    worksheet.cell(row=row_idx, column=col).value = value_widerstand
                    worksheet.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
                    # Temperatur
                    value_temperatur = row[df.columns.get_loc(f'{board_name} Temperatur')]
                    worksheet.cell(row=row_idx, column=col + 1).value = value_temperatur
                    worksheet.cell(row=row_idx, column=col + 1).alignment = Alignment(horizontal='center')

                    # Speichere die Zelladresse der Temperatur
                    temperatur_cell = worksheet.cell(row=row_idx, column=col + 1)
                    temperatur_cells.append(temperatur_cell.coordinate)

                    col += 2

                # Formel für gemittelte Temperatur
                if temperatur_cells:
                    formel = f"=ROUND(AVERAGE({','.join(temperatur_cells)}),2)"
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.value = formel
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # Falls keine Temperaturwerte vorhanden sind
                    worksheet.cell(row=row_idx, column=col).value = None

            # Spaltenbreiten anpassen
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

            # Rahmen hinzufügen
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border

            # Erste drei Zeilen grün einfärben
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            for row in range(1, 4):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = green_fill

        print(f'Daten erfolgreich in {dateiname} gespeichert.')

