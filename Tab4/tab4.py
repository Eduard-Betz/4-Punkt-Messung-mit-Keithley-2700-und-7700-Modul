# Datei: tab4.py

import tkinter as tk
from tkinter import ttk
from time import sleep

from global_vars import PlotAuswahl, TKBoardVariabeln  # Importiere den Plot-Auswahl-Manager

from Tab4.TKPlot import (
    plot_resistance_vs_temperature,
    plot_resistance_vs_avg_temperature,
    fit_and_store_line_data,
    fit_and_store_line_data_avg_temperature
)

from Tab4.AuswahlTK import open_widerstand_window
from Tab4.TKPlotControls import create_multiple_boards

# Debug-Flag definieren
debug_Tab4 = False  # Debug-Ausgaben sind standardmäßig deaktiviert

# Debug-Ausgabefunktion
def debug_print(*args, **kwargs):
    if debug_Tab4:
        print(*args, **kwargs)


def create_tab4(notebook):
    tab4 = ttk.Frame(notebook)
    notebook.add(tab4, text="Tab 4")

    # Frame für den ersten und zweiten Plot
    plot_frame1 = tk.Frame(tab4)
    plot_frame1.place(relx=0.01, relwidth=0.38, relheight=0.5)
    plot_frame2 = tk.Frame(tab4)
    plot_frame2.place(relx=0.01, rely=0.5, relwidth=0.38, relheight=0.5)

    # Notebook für die Boards innerhalb von tab4
    boards_notebook = ttk.Notebook(tab4)
    boards_notebook.place(relx=0.40, relwidth=0.4, relheight=1)  # Platzierung neben den Plotframes

    # Funktion zum Löschen der Plots und Zurücksetzen der Daten
    def clear_plots_and_data():
        # Lösche den Inhalt von plot_frame1 und plot_frame2
        for widget in plot_frame1.winfo_children():
            widget.destroy()  # Entferne alle Widgets innerhalb des Frames
        for widget in plot_frame2.winfo_children():
            widget.destroy()

        # Setze die Daten in PlotAuswahl auf None oder eine neutrale Einstellung zurück
        PlotAuswahl.reset_steigung()
        PlotAuswahl.reset_sinkend()

    # Funktion, um das Fenster für die Widerstandsauswahl zu öffnen und dann die Plots zu erstellen
    def open_and_plot():
        # Zuerst die Plots leeren und die Daten zurücksetzen
        clear_plots_and_data()

        # Erstelle ein neues Fenster für die Widerstandsauswahl
        window = tk.Toplevel()

        # Öffne das Fenster und pausiere die Ausführung bis das Fenster geschlossen ist
        open_widerstand_window(window)
        window.wait_window()  # Warten, bis das Fenster geschlossen ist

        # Erstelle die beiden Plots mit den Daten aus beiden Flanken
        plot_resistance_vs_temperature(plot_frame1, PlotAuswahl.get_steigung(), PlotAuswahl.get_sinkend())
        plot_resistance_vs_avg_temperature(plot_frame2, PlotAuswahl.get_steigung(), PlotAuswahl.get_sinkend())

    # Funktion, um die TK-Berechnung durchzuführen und die Boards anzuzeigen
    def TKBerechnen():
        # Führe die Berechnungen durch und aktualisiere die Plots
        fit_and_store_line_data(plot_frame1, PlotAuswahl.get_steigung(), PlotAuswahl.get_sinkend())
        fit_and_store_line_data_avg_temperature(plot_frame2, PlotAuswahl.get_steigung(), PlotAuswahl.get_sinkend())

        sleep(0.1)
        # Lösche die Tabs im boards_notebook, bevor neue erstellt werden
        for tab in boards_notebook.tabs():
            boards_notebook.forget(tab)

        # Erstellen der Board-Controls mit Tabs
        create_multiple_boards(boards_notebook)

    # Erstelle die Buttons
    button_positions = [
        # Dieser Button öffnet zuerst die Auswahl, löscht Plots, setzt Daten zurück und erstellt dann neue Plots
        ("Plot erstellen", 0, open_and_plot),
        ("TK Berechnen", 0.1, TKBerechnen),
        ("TK zu Exel Tabelle", 0.8, print_tk_data),
        ("Knopf 4", 0.9, None)
    ]

    buttons = create_buttons(tab4, button_positions)

    # Rückgabe des tab4
    return tab4

def create_buttons(tab, button_positions):
    buttons = []
    for text, rel_y, command in button_positions:
        button = tk.Button(tab, text=text, command=command)
        button.place(relx=1, rely=rel_y, anchor='ne', relwidth=0.2, relheight=0.1)
        buttons.append(button)
    return buttons


def print_tk_data():
    import pandas as pd
    from tkinter import filedialog
    import os
    from openpyxl.styles import Font, Alignment, Border, Side
    import openpyxl.utils
    from pprint import pprint  # Import für Debugging

    # Öffne einen Dialog, um den Speicherort der Excel-Datei zu wählen
    file_path = filedialog.asksaveasfilename(
        defaultextension='.xlsx',
        filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
        title='Excel-Datei speichern'
    )

    if not file_path:
        # Benutzer hat den Dialog abgebrochen
        return

    data = TKBoardVariabeln.board_variablen

    # Debug-Ausgabe hinzufügen
    debug_print("Inhalt von TKBoardVariabeln.board_variablen:")
    debug_print(data)

    # Listen zur Speicherung der Daten für die Tabellen
    normal_data_list = []
    avg_data_list = []

    # Daten extrahieren
    for board_nr, resistors in data.items():
        for resistor_nr, values in resistors.items():
            # Gemeinsame Daten
            row_common = {
                'Board Nr.': board_nr,
                'Probe': '',  # Diese Spalte bleibt leer
            }

            # Normal Daten
            if 'normal' in values:
                normal = values['normal']
                # Runden der Temperaturwerte auf zwei Dezimalstellen
                temp_steigend = normal.get('temp_bereich_steigend', (None, None))
                temp_sinkend = normal.get('temp_bereich_sinkend', (None, None))

                temp_steigend = tuple(round(t, 2) if isinstance(t, (int, float)) else t for t in temp_steigend)
                temp_sinkend = tuple(round(t, 2) if isinstance(t, (int, float)) else t for t in temp_sinkend)

                normal_row = row_common.copy()
                normal_row.update({
                    'TK Steigende Flanke': normal.get('steigung_steigend', ''),
                    'min temp Steigende Flanke': temp_steigend[0],
                    'max temp Steigende Flanke': temp_steigend[1],
                    'TK sinkende Flanke': normal.get('steigung_sinkend', ''),
                    'min temp sinkende Flanke': temp_sinkend[0],
                    'max temp sinkende Flanke': temp_sinkend[1],
                })
                normal_data_list.append(normal_row)

            # Avg Daten
            if 'avg' in values:
                avg = values['avg']
                # Runden der Temperaturwerte auf zwei Dezimalstellen
                temp_steigend = avg.get('temp_bereich_steigend', (None, None))
                temp_sinkend = avg.get('temp_bereich_sinkend', (None, None))

                temp_steigend = tuple(round(t, 2) if isinstance(t, (int, float)) else t for t in temp_steigend)
                temp_sinkend = tuple(round(t, 2) if isinstance(t, (int, float)) else t for t in temp_sinkend)

                avg_row = row_common.copy()
                avg_row.update({
                    'TK Steigende Flanke': avg.get('steigung_steigend', ''),
                    'min temp Steigende Flanke': temp_steigend[0],
                    'max temp Steigende Flanke': temp_steigend[1],
                    'TK sinkende Flanke': avg.get('steigung_sinkend', ''),
                    'min temp sinkende Flanke': temp_sinkend[0],
                    'max temp sinkende Flanke': temp_sinkend[1],
                })
                avg_data_list.append(avg_row)

    # Überprüfen, ob Daten in den Listen vorhanden sind
    if not normal_data_list and not avg_data_list:
        debug_print("Keine Daten zum Exportieren gefunden.")
        return

    # DataFrames erstellen
    columns = [
        'Board Nr.', 'Probe', 'TK Steigende Flanke', 'min temp Steigende Flanke', 'max temp Steigende Flanke',
        'TK sinkende Flanke', 'min temp sinkende Flanke', 'max temp sinkende Flanke'
    ]
    normal_df = pd.DataFrame(normal_data_list, columns=columns)
    avg_df = pd.DataFrame(avg_data_list, columns=columns)

    # Daten in Excel schreiben
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        normal_header = "TK mit Board Temperatur"
        avg_header = "TK mit gemittelte Temperatur über alle Boards"

        # Schreibe die normale Tabelle ab Zeile 3 (startrow=2 wegen nullbasierter Indizierung)
        normal_table_start_row = 2
        normal_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=normal_table_start_row, header=True)

        # Jetzt haben wir Daten geschrieben, wir können die workbook und worksheet erhalten
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Schreibe die Überschrift für die erste Tabelle in Zeile 1
        normal_header_row_excel = 1
        worksheet.cell(row=normal_header_row_excel, column=1).value = normal_header
        worksheet.merge_cells(start_row=normal_header_row_excel, start_column=1, end_row=normal_header_row_excel, end_column=len(columns))
        header_cell = worksheet.cell(row=normal_header_row_excel, column=1)
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal='center')

        # Berechne die Endzeile der normalen Tabelle
        normal_table_end_row = normal_table_start_row + len(normal_df)

        # Füge 10 leere Zeilen zwischen den Tabellen
        avg_header_row = normal_table_end_row + 10

        # Schreibe die Überschrift für die zweite Tabelle
        worksheet.cell(row=avg_header_row, column=1).value = avg_header
        worksheet.merge_cells(start_row=avg_header_row, start_column=1, end_row=avg_header_row, end_column=len(columns))
        header_cell = worksheet.cell(row=avg_header_row, column=1)
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal='center')

        # Schreibe die avg-Tabelle ab der berechneten Zeile (avg_header_row + 2)
        avg_table_start_row = avg_header_row + 2  # +2 um die Spaltenüberschriften einzuschließen
        avg_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=avg_table_start_row - 1, header=True)

        # Berechne die Endzeile der avg-Tabelle
        avg_table_end_row = avg_table_start_row + len(avg_df) - 1

        # Rahmen und Formatierungen anwenden
        # Definiere Rahmenstile
        thin_border = Side(border_style="thin", color="000000")
        thick_border = Side(border_style="medium", color="000000")

        def apply_table_border(ws, start_row, end_row, start_col, end_col):
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                for cell in row:
                    border_sides = {'left': thin_border, 'right': thin_border, 'top': thin_border, 'bottom': thin_border}

                    if cell.row == start_row:
                        border_sides['top'] = thick_border
                    if cell.row == end_row:
                        border_sides['bottom'] = thick_border
                    if cell.column == start_col:
                        border_sides['left'] = thick_border
                    if cell.column == end_col:
                        border_sides['right'] = thick_border

                    cell.border = Border(**border_sides)

        # Rahmen für die normale Tabelle anwenden
        apply_table_border(
            worksheet,
            normal_header_row_excel,
            normal_table_end_row + 1,  # +1 um die Überschrift einzuschließen
            1,
            len(columns)
        )

        # Rahmen für die avg-Tabelle anwenden (angepasst)
        apply_table_border(
            worksheet,
            avg_header_row,           # Startzeile des Rahmens (Überschrift der avg-Tabelle)
            avg_table_end_row + 1,    # Endzeile des Rahmens (+1 um die letzte Datenzeile einzuschließen)
            1,
            len(columns)
        )

        # Spaltenbreiten anpassen
        for i, column in enumerate(columns, 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

        # Ausrichtung für alle Zellen anpassen
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Erfolgsmeldung bleibt unverändert
    print(f"Excel-Datei wurde gespeichert unter: {file_path}")

